# -*- coding: utf-8 -*-
"""Post-generation validation for tracking plan xlsx (tracking-auto-import)."""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from tracking_utils import SKILL_ROOT, validate_template

# Optional AI attribute dictionary (same source as ai-tracking-validate)
_DICT_CANDIDATES = [
    Path(r"C:\Users\Administrator\AppData\Roaming\WPS 灵犀\serverdir\skills\ai-tracking-validate\attributes\属性字典.json"),
    Path(r"C:\Users\Administrator\AppData\Roaming\WPS 灵犀\serverdir\user_skills\ai-tracking-validate\attributes\属性字典.json"),
    SKILL_ROOT.parent / "ai-tracking-validate" / "attributes" / "属性字典.json",
]

AI_FORBIDDEN_PUBLIC_ATTRS = frozenset(
    {"func_version", "is_login", "member_identity", "entry_id", "position", "client"}
)

# Placeholder types — skip slash parity for these H-column values
VALUE_TYPE_PLACEHOLDERS = frozenset(
    {"string", "number", "int", "float", "bool", "boolean", "json", "array", "object"}
)

VAGUE_DESC_PATTERNS = (
    re.compile(r"功能展示"),
    re.compile(r"功能点击"),
    re.compile(r"场景名"),
    re.compile(r"等$"),
)

# attribute-rules.md §2.4.4 — AI 业务允许的事件名（以 ksheet 为准，此为校验兜底白名单）
AI_CANONICAL_EVENTS = frozenset(
    {
        "ai_show",
        "ai_click",
        "ai_open",
        "ai_clickrequest",
        "ai_createrequest",
        "ai_requestresult",
        "ai_resultuse",
        "ai_finishrequest",
        "ai_funcshow",
        "ai_funcclick",
        "ai_feedback",
        "ai_newfile",
        "ai_funcshow_duration",
        "ai_funcload_duration",
        "ai_funcresult",
        "ai_blacklist_show",
        "ai_blacklist_click",
        "ai_realnameshow",
        "ai_realnameclick",
    }
)

AI_MAINLINE_EVENTS = frozenset(
    {
        "ai_show",
        "ai_click",
        "ai_open",
        "ai_clickrequest",
        "ai_createrequest",
        "ai_requestresult",
        "ai_resultuse",
        "ai_finishrequest",
    }
)

# 增值/PDF 式命名误入 AI 方案：ai_{功能}_{show|click|...}
AI_PER_FUNC_EVENT_RE = re.compile(
    r"^ai_[a-z0-9_]+_(show|click|use|result|close|display|load|stay)$",
    re.IGNORECASE,
)

OTHER_BUSINESS_EVENT_RE = re.compile(
    r"^(pdf|vas|docer|photo)_[a-z0-9_]+",
    re.IGNORECASE,
)


def _load_attr_dict() -> dict:
    for path in _DICT_CANDIDATES:
        if path.is_file():
            with path.open(encoding="utf-8") as f:
                return json.load(f)
    return {}


def _split_enum(text: str | None) -> list[str]:
    if text is None:
        return []
    s = str(text).strip()
    if not s:
        return []
    return [p.strip() for p in s.split("/") if p.strip()]


def _is_ai_business(business: str) -> bool:
    b = (business or "").strip().lower()
    return b in ("ai", "ai业务")


def _validate_ai_events(
    events: set[str],
    func_id: str,
    errors: list[str],
) -> None:
    """Detect AI business rule mixing (§2.4)."""
    if not events:
        return

    for event in sorted(events):
        if OTHER_BUSINESS_EVENT_RE.match(event):
            errors.append(
                f"埋点方案: AI 业务出现他业务前缀事件 `{event}`，禁止混用 PDF/增值/稻壳/图片命名"
            )
            continue

        if event in AI_CANONICAL_EVENTS:
            continue

        if AI_PER_FUNC_EVENT_RE.match(event):
            errors.append(
                f"埋点方案: AI 业务禁止 `{event}`（混用增值/PDF 的 {{功能}}_show/click 命名）；"
                f"应使用规范固定事件名如 ai_show/ai_click/…，功能差异写在 function_code 等属性"
            )
            continue

        if func_id and func_id in event:
            errors.append(
                f"埋点方案: AI 业务事件 `{event}` 含功能标识 `{func_id}`，"
                f"功能标识不得拼进事件名（应写在 function_code/ai_app 等属性）"
            )
            continue

        errors.append(
            f"埋点方案: AI 业务事件 `{event}` 不在规范白名单（§2.4.4），"
            f"不得自行发明；须来自 ksheet 或用户明确要求新增"
        )

    # 仅有「功能标识_show/click」类命名、完全没有 AI 主线 → 业务规则混用模式
    has_per_func_only = events and not any(e in AI_MAINLINE_EVENTS for e in events)
    has_pdf_style = any(AI_PER_FUNC_EVENT_RE.match(e) for e in events)
    if has_per_func_only and has_pdf_style:
        errors.append(
            "埋点方案: AI 业务仅有 ai_{功能}_show/click 类事件、缺少 ai_show→ai_finishrequest 主线，"
            "疑似把 PDF/增值命名套进 AI 业务"
        )


def _read_basic_info(ws) -> dict[str, str]:
    info: dict[str, str] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        key = str(row[0]).strip()
        zh = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        code = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        info[key] = code or zh
    return info


def validate_plan(path: Path) -> list[str]:
    errors: list[str] = []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ["需要安装 openpyxl: pip install openpyxl"]

    try:
        validate_template(path)
    except Exception as e:
        return [f"模板结构校验失败: {e}"]

    attr_dict = _load_attr_dict()
    comp_values = set((attr_dict.get("comp") or {}).get("values", {}).keys())
    project_values = set((attr_dict.get("project") or {}).get("values", {}).keys())

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        basic = _read_basic_info(wb["基本信息"])
        business = basic.get("业务归属", "").lower()

        # --- 公共属性 Sheet ---
        pub = wb["公共属性"]
        pub_rows = list(pub.iter_rows(min_row=2, values_only=True))
        func_id = basic.get("功能名称", "")  # English id in col C row 功能名称
        if not func_id:
            for row in wb["基本信息"].iter_rows(min_row=3, values_only=True):
                if row and str(row[0]).strip() == "功能名称" and len(row) > 2 and row[2]:
                    func_id = str(row[2]).strip()
                    break

        for idx, row in enumerate(pub_rows, start=2):
            if not row or len(row) < 4:
                continue
            attr_name = str(row[1] or "").strip()
            attr_value = str(row[3] or "").strip()
            attr_desc = str(row[4] or "").strip() if len(row) > 4 else ""
            if not attr_name:
                continue

            if business in ("ai", "ai业务") and attr_name in AI_FORBIDDEN_PUBLIC_ATTRS:
                errors.append(f"公共属性 Sheet 第{idx}行: AI 业务禁止写入模板默认属性 `{attr_name}`")

            if "/" in attr_value or "/" in attr_desc:
                hv, iv = _split_enum(attr_value), _split_enum(attr_desc)
                if len(hv) != len(iv):
                    errors.append(
                        f"公共属性 Sheet 第{idx}行 `{attr_name}`: 属性值({len(hv)}段)与描述({len(iv)}段) `/` 分段不一致"
                    )

            if business in ("ai", "ai业务") and attr_name in ("comp", "project") and attr_value:
                known = comp_values if attr_name == "comp" else project_values
                if known and attr_value not in known:
                    errors.append(
                        f"公共属性 Sheet 第{idx}行: `{attr_name}` 值 `{attr_value}` 不在属性字典"
                    )
                if attr_value.startswith("ai_") and attr_value != "standalone":
                    errors.append(
                        f"公共属性 Sheet 第{idx}行: `{attr_name}` 误填功能标识 `{attr_value}`，独立应用应填 `standalone`"
                    )

            if business in ("ai", "ai业务") and attr_name == "ai_app":
                if attr_value == "standalone":
                    errors.append(f"公共属性 Sheet 第{idx}行: `ai_app` 不得为 `standalone`，应填功能标识")
                if func_id and attr_value and attr_value != func_id:
                    errors.append(
                        f"公共属性 Sheet 第{idx}行: `ai_app`=`{attr_value}` 与基本信息功能标识 `{func_id}` 不一致"
                    )

        # --- 埋点方案 Sheet (通用模型) ---
        plan = wb["埋点方案"]
        plan_rows = list(plan.iter_rows(min_row=3, values_only=True))
        is_common = (plan.max_column or 0) <= 11

        event_first_row: dict[str, int] = {}
        all_events: set[str] = set()
        for ridx, row in enumerate(plan_rows, start=3):
            if not row or len(row) < 9:
                continue
            event = str(row[2] or "").strip() if is_common else str(row[3] or "").strip()
            attr = str(row[4] or "").strip() if is_common else str(row[5] or "").strip()
            val = row[7] if is_common else row[8]
            desc = row[8] if is_common else row[9]
            val_s = str(val).strip() if val is not None else ""
            desc_s = str(desc).strip() if desc is not None else ""

            if event:
                all_events.add(event)
                if event in event_first_row:
                    errors.append(
                        f"埋点方案 第{ridx}行: 通用模型事件 `{event}` 重复（首次第{event_first_row[event]}行），应一事件一块"
                    )
                else:
                    event_first_row[event] = ridx

            if not attr or not val_s:
                continue

            if val_s.lower() in VALUE_TYPE_PLACEHOLDERS:
                continue

            if "/" in val_s or "/" in desc_s:
                hv, iv = _split_enum(val_s), _split_enum(desc_s)
                if len(hv) != len(iv):
                    errors.append(
                        f"埋点方案 第{ridx}行 事件={event or '?'} 属性={attr}: "
                        f"H列({len(hv)}段)与I列({len(iv)}段) `/` 分段不一致"
                    )
                elif attr == "func_name" and desc_s:
                    for pat in VAGUE_DESC_PATTERNS:
                        if pat.search(desc_s) and len(iv) < len(hv):
                            errors.append(
                                f"埋点方案 第{ridx}行 `func_name`: 描述含概括用语且未逐枚举对应"
                            )
                            break

            if business in ("ai", "ai业务") and attr in ("comp", "project") and val_s:
                known = comp_values if attr == "comp" else project_values
                if known and val_s not in known:
                    errors.append(f"埋点方案 第{ridx}行: `{attr}` 值 `{val_s}` 不在属性字典")
                if val_s.startswith("ai_") and val_s != "standalone":
                    errors.append(
                        f"埋点方案 第{ridx}行: `{attr}` 误填功能标识 `{val_s}`，独立应用应填 `standalone`"
                    )

            if business in ("ai", "ai业务") and attr == "ai_app" and val_s == "standalone":
                errors.append(f"埋点方案 第{ridx}行: `ai_app` 不得为 `standalone`")

        if _is_ai_business(business) and is_common:
            _validate_ai_events(all_events, func_id, errors)

    finally:
        wb.close()

    return errors


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate tracking plan xlsx before delivery")
    parser.add_argument("xlsx", type=Path, help="Path to generated xlsx")
    args = parser.parse_args()

    path = args.xlsx.resolve()
    if not path.is_file():
        print(f"文件不存在: {path}", file=sys.stderr)
        return 2

    errors = validate_plan(path)
    if errors:
        print(f"校验失败 ({len(errors)} 项): {path}")
        for i, err in enumerate(errors, 1):
            print(f"  {i}. {err}")
        return 1

    print(f"校验通过: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
