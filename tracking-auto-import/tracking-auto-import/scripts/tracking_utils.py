# -*- coding: utf-8 -*-
"""Shared helpers for tracking-auto-import (paths, auth, templates, delper push)."""
from __future__ import annotations

import json
import os
import platform
import subprocess
import sys
from pathlib import Path
from typing import Any

# tracking-auto-import/scripts/tracking_utils.py -> skill root
SKILL_ROOT = Path(__file__).resolve().parent.parent
SHARED_BIN = SKILL_ROOT.parent / "_shared" / "bin"
ASSETS_DIR = SKILL_ROOT / "assets"

TEMPLATE_COMMON = ASSETS_DIR / "通用模型模板.xlsx"
TEMPLATE_LOCATION = ASSETS_DIR / "位置模型模板.xlsx"

AUTH_PATH = SHARED_BIN / "auth.json"

EXE_BY_OS = {
    "Windows": "tracking_project_create_windows.exe",
    "Darwin": "tracking_project_create_darwin.exe",
    "Linux": "tracking_project_create_linux.exe",
}

EXPECTED_SHEETS = ("基本信息", "公共属性", "埋点方案")
PLAN_COLS = {"common": 11, "location": 12}


def load_auth() -> dict[str, str]:
    """Load uid and x-api-key; accepts legacy ``api-key`` field."""
    if not AUTH_PATH.is_file():
        raise FileNotFoundError(
            f"未找到认证配置: {AUTH_PATH}\n"
            "请复制 auth.json.example 为 auth.json 并填写 uid、x-api-key。"
        )
    with AUTH_PATH.open(encoding="utf-8") as f:
        data = json.load(f)
    uid = str(data.get("uid", "")).strip()
    api_key = str(data.get("x-api-key") or data.get("api-key") or "").strip()
    if not uid or not api_key:
        raise ValueError("auth.json 须包含非空的 uid 与 x-api-key（或兼容字段 api-key）")
    return {"uid": uid, "x-api-key": api_key}


def get_push_exe() -> Path:
    name = EXE_BY_OS.get(platform.system())
    if not name:
        raise OSError(f"不支持的操作系统: {platform.system()}")
    exe = SHARED_BIN / name
    if not exe.is_file():
        raise FileNotFoundError(f"未找到推送工具: {exe}")
    return exe


def resolve_template(model: str) -> Path:
    """model: 'common' | 'location' | '通用模型' | '位置模型'"""
    m = model.lower().strip()
    if m in ("common", "通用", "通用模型"):
        path = TEMPLATE_COMMON
    elif m in ("location", "位置", "位置模型"):
        path = TEMPLATE_LOCATION
    else:
        raise ValueError(f"未知模型类型: {model}")
    if not path.is_file():
        raise FileNotFoundError(f"模板不存在: {path}")
    return path


def validate_template(path: Path | str, model: str | None = None) -> dict[str, Any]:
    """Validate xlsx template structure; returns summary dict."""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError("需要安装 openpyxl: pip install openpyxl") from e

    path = Path(path)
    if not path.is_file():
        raise FileNotFoundError(path)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets = wb.sheetnames
        if list(sheets) != list(EXPECTED_SHEETS):
            raise ValueError(f"Sheet 应为 {EXPECTED_SHEETS}，实际为 {sheets}")

        plan = wb["埋点方案"]
        cols = plan.max_column or 0
        inferred = "location" if cols >= 12 else "common"
        if model:
            expected_cols = PLAN_COLS["location" if "位置" in model or model == "location" else "common"]
            if cols != expected_cols:
                raise ValueError(f"埋点方案列数应为 {expected_cols}，实际为 {cols}")

        return {
            "path": str(path),
            "size_bytes": path.stat().st_size,
            "sheets": sheets,
            "plan_cols": cols,
            "plan_max_row": plan.max_row,
            "model": inferred,
            "ok": True,
        }
    finally:
        wb.close()


def push_xlsx(xlsx_path: str | Path, timeout: int = 120) -> subprocess.CompletedProcess:
    """Push local xlsx to delper via bundled exe."""
    xlsx_path = Path(xlsx_path).resolve()
    if not xlsx_path.is_file():
        raise FileNotFoundError(f"xlsx 不存在: {xlsx_path}")
    if xlsx_path.suffix.lower() != ".xlsx":
        raise ValueError("仅支持 .xlsx 文件")

    auth = load_auth()
    exe = get_push_exe()
    cmd = [
        str(exe),
        "-xlsx",
        str(xlsx_path),
        "-uid",
        auth["uid"],
        "-x-api-key",
        auth["x-api-key"],
    ]
    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=timeout,
    )


def push_ok(result: subprocess.CompletedProcess) -> bool:
    out = (result.stdout or "") + (result.stderr or "")
    return result.returncode == 0 and "200 OK" in out
