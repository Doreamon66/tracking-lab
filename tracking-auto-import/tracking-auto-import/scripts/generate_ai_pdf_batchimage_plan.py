# -*- coding: utf-8 -*-
"""One-off generator: AI business tracking plan for PDF batch image PRD."""
from __future__ import annotations

import shutil
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, Side

sys.path.insert(0, str(Path(__file__).resolve().parent))
from tracking_utils import resolve_template

OUTPUT = Path(r"C:\Users\Administrator\ai_pdf_batchimage_埋点方案.xlsx")

OPEN_POSITION = [
    ("batchinsert_main", "批量插入主界面"),
    ("batchextract_main", "批量提取主界面"),
    ("batchinsert_paste_guide", "批量插入粘贴引导"),
    ("batchextract_copy_guide", "批量提取复制成功菜单"),
    ("tab_insert", "插入Tab入口"),
    ("tab_edit", "编辑Tab入口"),
    ("contextmenu_insert", "右键插入图片入口"),
    ("filepicker_imageprocess", "文件选择窗口图片处理模块"),
    ("contextmenu_extract", "右键提取图片入口"),
    ("tab_convert_extract", "转换Tab提取图片入口"),
]
OPEN_POSITION_H = "/".join(x[0] for x in OPEN_POSITION)
OPEN_POSITION_I = "/".join(x[1] for x in OPEN_POSITION)

FUNCTION_CODE = [
    ("photo_clarity", "一键清晰"),
    ("photo_contrast_enhance", "对比增强"),
    ("photo_bw_enhance", "黑白强化"),
    ("photo_gray_enhance", "灰度提升"),
    ("photo_soft_enhance", "画质柔和"),
    ("photo_remove_watermark", "去水印"),
    ("photo_remove_writing", "去字迹"),
    ("photo_remove_moire", "去摩尔纹"),
    ("photo_remove_blackedge", "去黑边"),
    ("photo_correct", "纠偏矫正"),
]
FUNCTION_CODE_H = "/".join(x[0] for x in FUNCTION_CODE)
FUNCTION_CODE_I = "/".join(x[1] for x in FUNCTION_CODE)

FUNC_NAME_SHOW = [
    ("batchinsert_ai_panel", "批量插入主界面AI能力区"),
    ("batchextract_ai_panel", "批量提取主界面AI能力区"),
    ("batchinsert_paste_guide", "批量插入粘贴引导"),
    ("batchextract_copy_guide", "批量提取复制成功引导"),
]
FUNC_NAME_SHOW_H = "/".join(x[0] for x in FUNC_NAME_SHOW)
FUNC_NAME_SHOW_I = "/".join(x[1] for x in FUNC_NAME_SHOW)

FUNC_NAME_CLICK = [
    ("view_mode_switch", "视图模式切换"),
    ("batch_apply_current", "批量应用到选中图片"),
    ("batch_apply_all", "批量应用到全部选中图片"),
    ("undo", "撤销"),
    ("redo", "还原"),
    ("hide_guide_once", "引导单次隐藏"),
    ("hide_guide_doc", "引导当前文档隐藏"),
]
FUNC_NAME_CLICK_H = "/".join(x[0] for x in FUNC_NAME_CLICK)
FUNC_NAME_CLICK_I = "/".join(x[1] for x in FUNC_NAME_CLICK)

INTENTION_CODE_H = "wps_pdf_batchimage_clarity/wps_pdf_batchimage_erase/wps_pdf_batchimage_correct"
INTENTION_CODE_I = "批量图片画质提升/wps_pdf批量图片AI消除/wps_pdf批量图片纠偏矫正"

PUBLIC_ATTRS = [
    ("公共属性", "comp", "WPS组件", "pdf", "pdf组件", "PDF组件内插件"),
    ("公共属性", "ai_type", "AI工具类型", "ai_shortcut_buttonrequest", "一键式快捷ai功能", ""),
    ("公共属性", "open_position", "AI入口/来源位置", OPEN_POSITION_H, OPEN_POSITION_I, "open_position需数据BP登记"),
    ("公共属性", "integritycheckvalue", "WPS文档唯一ID", "string", "WPS文档icv值", "能获取必报"),
    ("公共属性", "cloud_file_id", "云文档ID", "string", "云文档ID", "能获取必报"),
    ("公共属性", "intention_code", "计费意图编码", INTENTION_CODE_H, INTENTION_CODE_I, "需运营登记"),
]

COMMON_PUB_EVENT = [
    ("comp", "WPS组件", "公共属性", "pdf", "pdf组件", ""),
    ("ai_type", "AI工具类型", "公共属性", "ai_shortcut_buttonrequest", "一键式快捷ai功能", ""),
    ("open_position", "AI入口/来源位置", "公共属性", OPEN_POSITION_H, OPEN_POSITION_I, ""),
    ("integritycheckvalue", "WPS文档唯一ID", "公共属性", "string", "WPS文档icv值", ""),
    ("cloud_file_id", "云文档ID", "公共属性", "string", "云文档ID", ""),
]

EVENTS = [
    {
        "cat1": "批量图片AI",
        "cat2": "入口曝光",
        "code": "ai_show",
        "name": "入口展示",
        "trigger": "AI功能入口或AI能力区曝光时上报（批量插入/提取主界面AI区、粘贴/复制引导内AI入口等）",
        "attrs": COMMON_PUB_EVENT
        + [
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "按需；需运营登记function_code"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
            ("rp", "上报概率", "自定义属性", "bigint", "上报概率（如1000表示100%）", "按需"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "入口点击",
        "code": "ai_click",
        "name": "入口点击",
        "trigger": "用户点击AI功能入口时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "按需"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "功能打开",
        "code": "ai_open",
        "name": "功能打开",
        "trigger": "打开批量图片AI编辑/处理能力面板或进入AI批处理会话时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "必报"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
            ("track_id", "链路追踪ID", "自定义属性", "string", "track_id", "必报；ai_open生成并透传"),
            ("plug_version_name", "插件名", "自定义属性", "string", "插件名称", "能获取必报"),
            ("plug_version", "插件版本", "自定义属性", "string", "插件版本号", "能获取必报"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "发起请求点击",
        "code": "ai_clickrequest",
        "name": "点击发起请求",
        "trigger": "用户点击一键清晰/AI消除/纠偏等发起AI批处理请求时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "必报"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
            ("request_source", "请求发起来源", "自定义属性", "request/retry/fail_retry", "正常请求/重试/失败重试", "必报"),
            ("track_id", "链路追踪ID", "自定义属性", "string", "track_id", "必报"),
            ("action_id", "用户行为ID", "自定义属性", "string", "action_id", "必报；本事件生成并透传"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "客户端请求",
        "code": "ai_createrequest",
        "name": "客户端发起请求",
        "trigger": "客户端向AI网关发起批量图片处理请求时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("intention_code", "计费意图编码", "公共属性", INTENTION_CODE_H, INTENTION_CODE_I, "必报；需运营登记"),
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "必报"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
            ("request_source", "请求发起来源", "自定义属性", "request/retry/fail_retry", "正常请求/重试/失败重试", "必报"),
            ("request_id", "AI请求ID", "自定义属性", "string", "request_id", "必报；本事件生成并透传"),
            ("track_id", "链路追踪ID", "自定义属性", "string", "track_id", "必报"),
            ("action_id", "用户行为ID", "自定义属性", "string", "action_id", "必报"),
            ("ai_request_content", "AI请求内容", "自定义属性", "string", "AI请求内容摘要", "能获取必报"),
            ("image_count", "处理图片数", "自定义属性", "bigint", "本次批量处理的图片张数", "能获取必报"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "请求结果",
        "code": "ai_requestresult",
        "name": "AI请求返回结果",
        "trigger": "AI网关返回批量图片处理结果时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("intention_code", "计费意图编码", "公共属性", INTENTION_CODE_H, INTENTION_CODE_I, "必报"),
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "必报"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
            ("request_source", "请求发起来源", "自定义属性", "request/retry/fail_retry", "正常请求/重试/失败重试", "必报"),
            ("result", "请求结果", "自定义属性", "success/fail", "成功/失败", "必报"),
            ("is_show_success", "是否展示成功态", "自定义属性", "true/false", "是/否", "能获取必报"),
            ("product_name", "产品标识", "自定义属性", "wps-pdf-pc/wps-pdf-mac", "win客户端/Mac客户端", "必报"),
            ("request_id", "AI请求ID", "自定义属性", "string", "request_id", "必报"),
            ("duration", "耗时", "自定义属性", "bigint", "AI请求耗时(毫秒)", "必报"),
            ("response_content", "AI响应内容", "自定义属性", "string", "AI响应内容摘要", "能获取必报"),
            ("fail_reason", "失败原因", "自定义属性", "string", "失败原因描述", "失败时上报"),
            ("success_count", "成功张数", "自定义属性", "bigint", "批量处理成功图片张数", "能获取必报"),
            ("fail_count", "失败张数", "自定义属性", "bigint", "批量处理失败图片张数", "能获取必报"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "结果决策",
        "code": "ai_resultuse",
        "name": "返回结果决策使用",
        "trigger": "用户对AI批处理结果执行应用/撤销/还原/放弃时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("answer_use", "结果使用决策", "自定义属性", "use/abandon/undo/redo", "应用/放弃/撤销/还原", "必报；answer_use需登记AI决策类别字典"),
            ("request_id", "AI请求ID", "自定义属性", "string", "request_id", "必报"),
            ("track_id", "链路追踪ID", "自定义属性", "string", "track_id", "必报"),
            ("action_id", "用户行为ID", "自定义属性", "string", "action_id", "必报"),
        ],
    },
    {
        "cat1": "批量图片AI",
        "cat2": "请求结束",
        "code": "ai_finishrequest",
        "name": "AI当次结束请求",
        "trigger": "单次AI批量处理链路结束（成功/失败/中止）时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("function_code", "请求功能", "自定义属性", FUNCTION_CODE_H, FUNCTION_CODE_I, "必报"),
            ("function_name", "功能名称", "自定义属性", "string", "AI功能中文名称", "按需"),
            ("track_id", "链路追踪ID", "自定义属性", "string", "track_id", "必报"),
            ("action_id", "用户行为ID", "自定义属性", "string", "action_id", "必报"),
            ("request_id", "AI请求ID", "自定义属性", "string", "request_id", "能获取必报"),
            ("scene", "结束场景", "自定义属性", "finish/abort", "正常结束/人为中止", "必报"),
            ("total_duration", "总耗时", "自定义属性", "bigint", "单次AI批处理总耗时(毫秒)", "必报"),
            ("discontinue_type", "中断类型", "自定义属性", "artificial/regular/other", "手动终止/规则拦截/其他", "必报"),
            ("discontinue_reason", "中断原因", "自定义属性", "string", "中断原因描述", "能获取必报"),
        ],
    },
    {
        "cat1": "批量图片插件",
        "cat2": "功能展示",
        "code": "ai_funcshow",
        "name": "AI功能展示",
        "trigger": "批量插入/提取插件内AI相关模块或引导曝光时上报",
        "attrs": COMMON_PUB_EVENT
        + [
            ("func_name", "功能场景名", "自定义属性", FUNC_NAME_SHOW_H, FUNC_NAME_SHOW_I, "合并上报；区分插件场景"),
            ("page_name", "页面名称", "自定义属性", "batchinsert/batchextract", "批量插入/批量提取", "按需"),
        ],
    },
    {
        "cat1": "批量图片插件",
        "cat2": "功能点击",
        "code": "ai_funcclick",
        "name": "AI功能点击",
        "trigger": "批量图片插件内非AI请求类交互点击时上报（视图切换、撤销还原、引导隐藏等）",
        "attrs": COMMON_PUB_EVENT
        + [
            ("func_name", "功能场景名", "自定义属性", FUNC_NAME_CLICK_H, FUNC_NAME_CLICK_I, "合并上报"),
            ("page_name", "页面名称", "自定义属性", "batchinsert/batchextract", "批量插入/批量提取", "按需"),
        ],
    },
]


def _style_row(ws, row: int) -> None:
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font = Font(name="微软雅黑", size=10.5)
    align = Alignment(vertical="center", wrap_text=True)
    for col in range(1, 12):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.alignment = align
        cell.border = border


def _clear_sheet(ws, start_row: int = 3) -> None:
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row >= start_row:
            ws.unmerge_cells(str(merged))
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def main() -> Path:
    template = resolve_template("common")
    if OUTPUT.exists():
        OUTPUT.unlink()
    shutil.copy(template, OUTPUT)

    wb = load_workbook(OUTPUT)
    basic = wb["基本信息"]
    pub = wb["公共属性"]
    plan = wb["埋点方案"]

    basic["B3"], basic["C3"] = "WPS Office", "wps"
    basic["B4"], basic["C4"] = "PC端", "pc"
    basic["B5"], basic["C5"] = "AI业务", "ai"
    basic["B6"], basic["C6"] = "AI_PDF批量图片", "ai_pdf_batchimage"
    basic["B7"] = "【PC PDF】批量插入和导出图片升级"
    basic["C7"] = "https://365.kdocs.cn/l/chxC46Ewj3c4"

    _clear_sheet(pub, 2)
    for i, row in enumerate(PUBLIC_ATTRS, start=2):
        for j, val in enumerate(row, start=1):
            pub.cell(i, j, val)
        _style_row(pub, i)

    _clear_sheet(plan, 3)
    row = 3
    merge_ranges = []

    for ev in EVENTS:
        start = row
        plan.cell(row, 1, ev["cat1"])
        plan.cell(row, 2, ev["cat2"])
        plan.cell(row, 3, ev["code"])
        plan.cell(row, 4, ev["name"])
        plan.cell(row, 11, ev["trigger"])
        _style_row(plan, row)
        row += 1
        for attr in ev["attrs"]:
            plan.cell(row, 5, attr[0])
            plan.cell(row, 6, attr[1])
            plan.cell(row, 7, attr[2])
            plan.cell(row, 8, attr[3])
            plan.cell(row, 9, attr[4])
            plan.cell(row, 10, attr[5])
            _style_row(plan, row)
            row += 1
        end = row - 1
        merge_ranges.append((start, end, 1))
        merge_ranges.append((start, end, 2))
        merge_ranges.append((start, end, 3))
        merge_ranges.append((start, end, 4))
        merge_ranges.append((start, end, 11))

    for start, end, col in merge_ranges:
        if start < end:
            plan.merge_cells(start_row=start, start_column=col, end_row=end, end_column=col)

    wb.save(OUTPUT)
    wb.close()
    return OUTPUT


if __name__ == "__main__":
    out = main()
    print(out)
